# -*- coding: cp1252 -*-
#This tutorial was built by me, Farzain! You can ask me questions or troll me on Twitter (@farzatv)

#First we need to import requests. Installing this is a bit tricky. I included a step by step process on how to get requests in readme.txt which is included in the file along with this program.
import requests
import os
import requestsEngine
import spreadsheetEngine
from APIKeyFile import APIKey
import staticRequests

def spreadsheetExistCheck():#       Para revisar si hay una Spreadsheet existente en el folder
    controlSpeadsheet = os.path.exists('RankedData.xlsx')
    return controlSpeadsheet    

def spreadsheetGraber():# Para seleccionar la spreadsheet
    global wb
    global ws
    from openpyxl import load_workbook
    wb = load_workbook('RankedData.xlsx')

def getRankedSoloMatchlist():
    global summonerRankedSoloMatchlist
    summonerRankedSoloMatchlist = requestsEngine.requestRankedSoloMatchlist(summonerData, APIKey)

def getRankedMatchIDs():
    global RankedMatchIDs
    RankedMatchIDs = []
    i = 0
    while i < summonerRankedSoloMatchlist['totalGames']:
        RankedMatchIDs.append(summonerRankedSoloMatchlist['matches'][i]['matchId'])
        i += 1

def spreadsheetDataGraber():
    spreadsheetGraber()
    ws = wb["Summoner Info"]
    ID = (str)(ws['B2'].value)
    summonerName = str.lower((str)(ws['B3'].value))
    summonerData = {}
    summonerData[summonerName] = {}
    summonerData[summonerName]['id'] = (str)(ws['B2'].value)
    summonerData[summonerName]['name'] = (str)(ws['B3'].value)
    summonerData[summonerName]['profileIconId'] = (str)(ws['B5'].value)
    summonerData[summonerName]['revisionDate'] = (ws['A1'].value)
    summonerData[summonerName]['summonerLevel'] = (str)(ws['B6'].value)
    summonerData[summonerName]['region'] = str.lower((str)(ws['B4'].value))
    return summonerData

def main():#                    ---El programa---

    print "\nWelcome Summoner!!"
    name = (str)(raw_input('\nType your Summoner Name here and DO NOT INCLUDE ANY SPACES: '))
    nameMin = str.lower(name)
    summonerName = name
    
    spreadsheetExistCheck()# Si existe spreadsheet (datos) verifica cambios
    if spreadsheetExistCheck():
        summonerData = spreadsheetDataGraber()
        if nameMin == summonerData.keys()[0]:# Compara los datos con los que "inicia sesion" con los guadados
            print "Hello " + name

        else:
            summonerName = nameMin
            print "\nYour summoner Data might not be updated."
            print "\nPleas enter your region to get started"
            print "Type in one of the following regions or else the program wont work correctly:\n"
            print "NA EUW EUNE LAN BR KR LAS OCE TR RU PBE\n"
            region = str.lower((str)(raw_input('Type in one of the regions above: ')))
            summonerData = requestsEngine.requestSummonerData(region, summonerName, APIKey)
            spreadsheetEngine.spreadsheetUpdater(summonerData, APIKey)
            print "\nYour Ranked Statistics Spreadsheet is up to date\n"
               
    else:#   Obtener los los datos del invocador
        print "\nEnter your region to get started"
        print "Type in one of the following regions or else the program wont work correctly:\n"
        print "NA EUW EUNE LAN BR KR LAS OCE TR RU PBE\n"
    
        region = str.lower((str)(raw_input('Type in one of the regions above: ')))
        summonerData = requestsEngine.requestSummonerData(region, summonerName, APIKey)
        summonerData[summonerName]['region'] = region

    #   Establecido el ID. Preguntar que hacer
    control = 0
    while control != 6:
        print "\nWhat do you want to do?"
        print "\nPruebas escribe 0"
        print "Type 1 for Rankded Stats"
        print "Type 2 for last match result"
        print "Type 3 if you want to creat your Ranked Statistics Spreadsheet"
        print "Type 4 to update your Ranked Statistics Spreadsheet"
        print "Type 5 to get your ranked list"
        print "Type 6 for Exit\n"
        control = input()
        
        if control == 1:#       Para imprimir datos de Ranked          
            summonerRankedData = requestsEngine.requestRankedData(summonerData, APIKey)
            ID = (str)(summonerData[summonerName]['id'])
            summonerLP = summonerRankedData[ID][0]['entries'][0]['leaguePoints']# Se tiene que convertir a str
            summonerLP = str(summonerLP)
            print "Your current Tier is: " + summonerRankedData[ID][0]['tier']
            print "Your current Division is: " + summonerRankedData[ID][0]['entries'][0]['division']
            print "And you currently have: " + summonerLP + " League Points"
            print "\nDone!\n"

        elif control == 2:#     Imprimir resultado del ultimo match
            summonerRecentGames = requestsEngine.requestRecentGames(summonerData, APIKey)
            if summonerRecentGames['games'][0]['stats']['win']:
                print "\nYou Win\n"
            else:
                print "\nYou Loss\n"

        elif control == 3:#     Crear Ranked Spreadsheet de Invocador
            if spreadsheetExistCheck():#        Revisa si exite ya un archivo primero
                print "\nThere is already a Ranked Statistics Spreadsheet\n"
            else:#                              Si no hay archivo, procede
                spreadsheetEngine.createSpreadsheet(summonerData, APIKey)
                spreadsheetEngine.spreadsheetUpdater(summonerData, APIKey)
                
        elif control == 4:#      Actualizar Ranked Spreadsheet
            if spreadsheetExistCheck():#        Revisa si existe spreadsheet que actualizar
                spreadsheetGraber()
                ws = wb["Summoner Info"]
                controlDe4 = (ws['A1'].value)#  Valor del ultimo cambio en la spreadsheet actual
                summonerName = summonerData.keys()[0]
                
                if controlDe4 == summonerData[summonerName]['revisionDate']:#     Valida si el Revision Date es el mismo
                    print "\nYour Ranked Statistics Spreadsheet is up to date\n"
                else:#      Si el Revision Date no es el mismo se actulizará la spreadsheet
                    spreadsheetEngine.spreadsheetUpdater(summonerData, APIKey)
                                                            
            else:#           Si no hay spreadsheet te ofrece crear una y se actualiza
                print "\nThere is no Ranked Statistics Spreadsheet. Would you like to create one now?"
                controlDe4 = (str)(raw_input("Write y for Yes or n for No\n"))
                if controlDe4 == "y":
                    spreadsheetEngine.createSpreadsheet(summonerData, APIKey)
                    spreadsheetEngine.spreadsheetUpdater(summonerData, APIKey)    
                    
                else:#                          Si dice que no, no hace nada
                    print "Ok"
                
        elif control == 5:# Llamar rankedMatchList
            getRankedSoloMatchlist()
            getRankedMatchIDs()
            
        elif control == 6:#     Para salir
            print "Ok. See ya"
            break

        elif control == 0:# Para pruebas
            from openpyxl.styles import Style, fills, PatternFill, Color
            from openpyxl.styles import colors
            spreadsheetEngine.createSpreadsheet(summonerData, APIKey)
            spreadsheetGraber()
            ws = wb["Summoner Info"]
            celdas = ws.range('A2:A6')
            celdas.style = Style(fill=PatternFill(patternType='solid', fgColor=Color('FFFF0000')))
            wb.save('RankedData.xlsx')
            

#This starts my program!
if __name__ == "__main__":
    main()

