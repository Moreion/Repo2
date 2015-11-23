#  -- Funciones para crear y modificar la Ranked Statistics Spreadsheet --

from openpyxl import Workbook
from RiotAPIApp import *

def createSpreadsheet(summonerData, APIKey):#       Para crear archivo xlsx con info de Invocador  ***Faltan cosas
        wb = Workbook()
        ws = wb.active#         Seleccionas la hoja activa por default se crea una
       #     Para crear la primer hoja que tendra datos generales del Invocador
        ws.title = "Summoner Info"
        ws.row_dimensions[1].hidden = True
        ws.column_dimensions['A'].width = 20
        ws['A1'] = 0# Agregamos el Revision Date para validar actualizaciones
        ws['A2'] = "Summoner ID:"
        ws['A3'] = "Summoner Name:"
        ws['A4'] = "Region:"
        ws['A5'] = "Icon:"
        ws['A6'] = "Level:"
        ws['A7'] = "Rank/Division:"
        ws['A8'] = "League Points:"
        ws['A9'] = "Win/Loss:"
        ws['A10'] = "KDA:"
        ws['A11'] = "Average Creep Score:"
        ws['D3'] = "My last 5 Games"
        ws['D4'] = "Type"
        ws['E4'] = "Result"
        ws['F4'] = "Champ"
        ws['G4'] = "KDA"
        ws['H4'] = "CS/min"
        wb.save('RankedData.xlsx')
        print "\nSpreadsheet created\n"

def spreadsheetUpdater(summonerData, APIKey):# Actualiza el spreadsheet *****Faltan muchas cosas
        summonerName = summonerData.keys()[0]
        ID = (str)(summonerData[summonerName]['id'])
        region = summonerData[summonerName]['region']
        summonerRankedData = requestsEngine.requestRankedData(summonerData, APIKey)
        from openpyxl import load_workbook
        # Actualizar la hoja de Summoner Info
        wb = load_workbook('RankedData.xlsx')
        ws = wb["Summoner Info"]
        ws['B2'] = summonerData[summonerName]['id']
        ws['B3'] = summonerData[summonerName]['name']
        ws['B4'] = str.capitalize(summonerData[summonerName]['region'])
        ws['B5'] = summonerData[summonerName]['profileIconId']
        ws['B6'] = summonerData[summonerName]['summonerLevel']
        ws['B7'] = summonerRankedData[ID][0]['tier'] + " " + summonerRankedData[ID][0]['entries'][0]['division']
        ws['B8'] = summonerRankedData[ID][0]['entries'][0]['leaguePoints']

        # Actualizar los ultimos 5 matches
        summonerRecentGames = requestsEngine.requestRecentGames(summonerData, APIKey)
        i = 0
        a = 5
        while i < 5:#Loop para llenar los datos de los 5 matches **** No me deja definir variables dentro del loop ** No puedo sumar para CS
                a = str(a)
                ws['D'+ a] = summonerRecentGames['games'][i]['subType']
                if summonerRecentGames['games'][i]['stats']['win']:
                        ws['E'+ a] = "Win"
                else:
                        ws['E' + a] = "Loss"
                ws['F' + a] = summonerRecentGames['games'][i]['championId']
                
                # Para poner las kills, deaths y assists primero hay que buscar si el Key existe en el dictionary
                if 'championsKilled' in summonerRecentGames['games'][i]['stats']:
                        kills = (str)(summonerRecentGames['games'][i]['stats']['championsKilled'])
                else:
                        kills = "0"
                if 'numDeaths' in summonerRecentGames['games'][i]['stats']:
                        deaths = (str)(summonerRecentGames['games'][i]['stats']['numDeaths'])
                else:
                        deaths = "0"
                if 'assists' in summonerRecentGames['games'][i]['stats']:
                        assists = (str)(summonerRecentGames['games'][i]['stats']['assists'])
                else:
                        assists = "0"
                ws['G' + a] = kills + "/" + deaths + "/" + assists

                # Para poner CS primero se busan y se suman los cs de jungla (tuya y enemiga) y minions
                if 'neutralMinionsKilled' in summonerRecentGames['games'][i]['stats']:
                        jungla = summonerRecentGames['games'][i]['stats']['neutralMinionsKilled']
                else:
                        jungla = 0
                if 'minionsKilled' in summonerRecentGames['games'][i]['stats']:
                        minions = summonerRecentGames['games'][i]['stats']['minionsKilled']
                else:
                        minions = 0
                CS = jungla + minions
                # Calcular los minutos jugados
                if 'timePlayed' in summonerRecentGames['games'][i]['stats']:
                        timeMin = summonerRecentGames['games'][i]['stats']['timePlayed'] / 60.0
                else:
                        timeMin = 0
                ws['H' + a] = format((CS / timeMin), '.2f')
                
                a = int(a)
                i += 1
                a += 1
        
        # Se cambia el revisionDate para validar futuras actualizaciones
        wb["Summoner Info"]['A1'] = summonerData[summonerName]['revisionDate']
        # Se guardan los cambios
        wb.save('RankedData.xlsx')
        print "\nYour Ranked Statistics Spreadsheet has been updated\n"
